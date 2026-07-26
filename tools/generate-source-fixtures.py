#!/usr/bin/env python3
"""Regenerate minimal, source-derived adapter fixtures.

The script intentionally copies only headers, notes, and a handful of observed
rows from each NJ DOE artifact. It never synthesizes data values and never
commits a complete upstream dataset.
"""

from __future__ import annotations

import hashlib
import io
import json
from pathlib import Path
import shutil
import tempfile
import urllib.request
import zipfile

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
FIXTURES = ROOT / "inst" / "extdata" / "test-fixtures" / "source-adapters"
USER_AGENT = "njschooldata fixture generator"

SOURCES = {
    "enrollment": "https://www.nj.gov/education/doedata/enr/enr20/enrollment_1920.zip",
    "assessment": (
        "https://www.nj.gov/education/assessment/results/reports/2425/spring/"
        "ELA04%20NJSLA%20DATA%202024-25.xlsx"
    ),
    "spr": (
        "https://www.nj.gov/education/sprreports/download/DataFiles/2024-2025/"
        "Database_DistrictStateDetail.xlsx"
    ),
    "finance": "https://www.nj.gov/education/guide/docs/2025/TGES2025_Zipped.zip",
}


def sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def download(url: str) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(request, timeout=1200) as response:
        return response.read()


def copy_rows(source, target, rows: list[int]) -> None:
    for out_row, source_row in enumerate(rows, start=1):
        for column in range(1, source.max_column + 1):
            target.cell(out_row, column, source.cell(source_row, column).value)


def workbook_bytes(workbook: Workbook) -> bytes:
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def enrollment_fixture(source: bytes) -> bytes:
    with zipfile.ZipFile(io.BytesIO(source)) as archive:
        member = next(name for name in archive.namelist() if name.lower().endswith(".xlsx"))
        workbook = load_workbook(io.BytesIO(archive.read(member)), data_only=True)

    fixture = Workbook()
    fixture.remove(fixture.active)
    for requested in ("State", "District", "School"):
        sheet = next(ws for ws in workbook.worksheets if ws.title.strip() == requested)
        target = fixture.create_sheet(sheet.title)
        selected = []
        if requested == "State":
            headers = [sheet.cell(3, column).value for column in range(1, sheet.max_column + 1)]
            grade_column = headers.index("Grade") + 1
            selected = [
                row for row in range(4, sheet.max_row + 1)
                if sheet.cell(row, grade_column).value in ("Eight Grade", "All Grades")
            ][:2]
        else:
            selected = [
                row for row in range(4, sheet.max_row + 1)
                if any(sheet.cell(row, column).value == ">95"
                       for column in range(1, sheet.max_column + 1))
            ][:2]
        if not selected:
            selected = list(range(4, min(sheet.max_row, 5) + 1))
        copy_rows(sheet, target, [1, 2, 3, *selected])

    xlsx = workbook_bytes(fixture)
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("enrollment_1920_fixture.xlsx", xlsx)
    return output.getvalue()


def assessment_fixture(source: bytes) -> bytes:
    workbook = load_workbook(io.BytesIO(source), data_only=True, read_only=True)
    sheet = workbook.active
    fixture = Workbook()
    target = fixture.active
    target.title = sheet.title
    data_rows = list(range(4, min(sheet.max_row - 2, 8) + 1))
    copy_rows(sheet, target, [1, 2, 3, *data_rows, sheet.max_row - 1, sheet.max_row])
    return workbook_bytes(fixture)


def spr_fixture(source: bytes) -> bytes:
    workbook = load_workbook(io.BytesIO(source), data_only=True, read_only=True)
    sheet = workbook["ChronicAbsenteeismStudentGroup"]
    fixture = Workbook()
    target = fixture.active
    target.title = sheet.title
    copy_rows(sheet, target, list(range(1, min(sheet.max_row, 9) + 1)))
    return workbook_bytes(fixture)


def finance_fixture(source: bytes) -> bytes:
    with zipfile.ZipFile(io.BytesIO(source)) as archive:
        members = [name for name in archive.namelist() if name.lower().endswith(".csv")]
        preferred = [name for name in members if Path(name).stem.upper() == "CSG1"]
        member = (preferred or members)[0]
        lines = archive.read(member).splitlines(keepends=True)
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(member, b"".join(lines[:10]))
    return output.getvalue()


def main() -> None:
    FIXTURES.mkdir(parents=True, exist_ok=True)
    builders = {
        "enrollment": ("enrollment-2020.zip", enrollment_fixture),
        "assessment": ("njsla-ela04-2025.xlsx", assessment_fixture),
        "spr": ("spr-district-2025.xlsx", spr_fixture),
        "finance": ("tges-2025.zip", finance_fixture),
    }
    provenance = {}
    spr_cache = (
        Path.home() / "Library" / "Caches" / "org.R-project.R" / "R" /
        "njschooldata" / "spr-workbooks" / "SPR_district_2025.xlsx"
    )
    for family, (filename, builder) in builders.items():
        if family == "spr" and spr_cache.exists():
            source = spr_cache.read_bytes()
            retrieval = "existing validated njschooldata workbook cache"
        else:
            source = download(SOURCES[family])
            retrieval = "direct HTTPS download"
        fixture = builder(source)
        path = FIXTURES / filename
        path.write_bytes(fixture)
        provenance[family] = {
            "source_url": SOURCES[family],
            "source_sha256": sha256(source),
            "fixture": filename,
            "fixture_sha256": sha256(fixture),
            "retrieval": retrieval,
            "extraction": "headers/preamble plus minimal representative observed rows",
        }

    (FIXTURES / "provenance.json").write_text(
        json.dumps(provenance, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )


if __name__ == "__main__":
    main()
